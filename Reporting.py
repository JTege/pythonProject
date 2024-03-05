# Automation

import openpyxl as jams
file = jams.load_workbook('Report.xlsx')
sheet = file['Sheet1']
cell = sheet['a1']  # coordinates of a cell
# print(cell.value)
# print(sheet.max_row)

for row in range(2, sheet.max_row + 1):  # rows
    cell = sheet.cell(row, 2)  # column 2
    # print(cell.value)
    average_value = cell.value / 12
    average_column = sheet.cell(row, 15)
    average_column.value = average_value

# print(sheet.max_column)

file.save('Report_v2.xlsx')
